import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import glob
import xlwings as xw
import zipfile

# 建立名称对应关系字典
name_dict = {}
name_dict['polyamide-6(volgamid 25)'] = '25'
name_dict['polyamide-6(volgamid 27)'] = '27'
name_dict['polyamide-6(volgamid 34F)'] = '34F'
name_dict['polyamide-6(volgamid 24)'] = '24'
name_dict['polyamide-6(volgamid 24SD)'] = '24 SD'
name_dict['polyamide-6(volgamid 32)'] = '32'
name_dict['polyamide-6(volgamid 34)'] = '34'
name_dict['polyamide-6(volgamid 36)'] = '36'
name_dict['polyamide-6(volgamid 40)'] = '40'
name_dict['CAPROLACTAM'] = 'Caprolactam'
name_dict['Nylon 6 dipped NTCF'] = 'NYLON 6 Dipped NTCF'
name_dict['在途物资-materials in transit'] = 'Inventories (tn)'
name_dict['库存商品-Commodity Stocks'] = 'Inventories (tn)'
name_dict['其他非主营'] = 'others '
###################################
# name_dict['Polyamide-6 Recycled'] = 'Polyamide light-stabilized yarn 93.5 tex'
# name_dict['Polyamide-6 Recycled'] = 'Polyamide not the  thermostabilized yarn 187tex'
# name_dict['Polyamide-6 Recycled'] = 'yarn 144tex'
# name_dict['Polyamide-6 Recycled'] = 'yarn 94tex'
# name_dict['Polyamide-6 Recycled'] = 'Polyamide fiber 1 tex'
# name_dict['Polyamide-6 Recycled'] = 'Polyamide fiber 0.48 tex'
# name_dict['Polyamide-6 Recycled'] = 'Polyamide fiber 0.68 tex'
# name_dict['Polyamide-6 Recycled'] = 'Polyamide fiber 0.33 tex'
# name_dict['Polyamide-6 Recycled'] = 'fish net'
# name_dict['Polyamide-6 Recycled'] = 'NYLON 6 Dipped NTCF'
# name_dict['Polyamide-6 Recycled'] = 'PA66TNC103'


def is_valid_xlsx(file_path):
    try:
        with zipfile.ZipFile(file_path) as z:
            return '[Content_Types].xml' in z.namelist()
    except zipfile.BadZipFile:
        return False
def find_excel_files_with_keyword(folder_path, keyword):
    # 构建搜索路径，匹配所有 Excel 文件（支持 .xlsx 和 .xls）
    search_pattern = os.path.join(folder_path, f"{keyword}*.xls*")
    # 使用 glob 查找匹配的文件
    matching_files = glob.glob(search_pattern)
    
    return matching_files[0]
def handle_transit_table(transit_table):
    """
    处理在途货物余额表，提取特定信息并生成产品列表。
    
    该函数读取一个Excel文件，遍历表格中的数据，寻找和提取特定产品的信息，
    包括产品名称、数量、价格等，并将这些信息存储在一个全局列表product_list中。
    """
    try:
        # 如果是 .xls 文件，先转成 .xlsx
        if transit_table.endswith(".xls") and not transit_table.endswith(".xlsx"):
            df_html = pd.read_html(transit_table)[0]
            old_file = transit_table
            transit_table = transit_table[:-4] + ".xlsx"
            df_html.to_excel(transit_table, index=False)
            # 删除原来的 .xls 文件
            os.remove(old_file)

        # 正常读取 Excel 文件
        if is_valid_xlsx(transit_table):
            df = pd.read_excel(transit_table, sheet_name=0, header=None, engine='openpyxl')
        else:
            df = pd.read_excel(transit_table, sheet_name=0, header=None, engine='xlrd')

    except Exception as e:
        print("无法处理该文件:", e)
        raise

    # 找到并记录月份
    global curr_month 
    global new_tonns_name
    # 遍历A列，找到包含“日期”的行号
    for index, value in enumerate(df[0]):
        if pd.notna(value) and '日期' in str(value):
            tmp = df.iloc[index, 1]
            cur_date = datetime.strptime(tmp, '%Y-%m-%d')
            curr_month = cur_date.month
            print('month:{}'.format(curr_month))
            new_tonns_name = f'tonns of good{cur_date.year}.xlsx'
            break

    product_list = []
    other = {}
    other['name'] = 'other'
    other['tonns_name'] = 'others '
    other['month'] = curr_month
    other['quantity'] = 0
    other['amount'] = 0

    # 遍历表格，提取产品信息
    for index, value in enumerate(df[1]):
        if index < 5: continue

        if pd.notna(value):
            exist = False

            for key in name_dict.keys():
                if key in str(value):
                    exist = True
                    break

            if True == exist:
                tmp_dic = {}
                tmp_dic['name'] = value

                if pd.notna(df.iloc[index, 7]):
                    tmp_dic['quantity'] = float(df.iloc[index, 7])
                else:
                    tmp_dic['quantity'] = 0

                if pd.notna(df.iloc[index, 8]):
                    tmp_dic['amount'] = float(df.iloc[index, 8])
                else:
                    tmp_dic['amount'] = 0
                
                tmp_dic['month'] = curr_month
                tmp_dic['tonns_name'] = name_dict[value]

            
                product_list.append(tmp_dic)
            else:
                other['amount'] += float(df.iloc[index, 8])
        
    product_list.append(other)
    return product_list
def handle_outbound_summary(outbound_table):
    # 读取出库汇总表
    if is_valid_xlsx(outbound_table):
        df = pd.read_excel(outbound_table, sheet_name=0, header=None, engine='openpyxl')
    else:
        df = pd.read_excel(outbound_table, sheet_name=0, header=None, engine='xlrd')    

    sale_list = []
    pattern = r'volgamid (\w+)\)'
    other = {}
    other['name'] = 'other'
    other['tonns_name'] = 'others '
    other['quantity'] = 0
    other['amount'] = 0

    # 遍历第一行，找到'数量'和'金额'所在的列索引
    quantity_column_index = None
    amount_column_index = None
    for index, value in enumerate(df.iloc[0]):
        if value == '数量':
            quantity_column_index = index
        elif value == '金额':
            amount_column_index = index
            break

    # 遍历表格，提取产品信息
    for index, value in enumerate(df[7]):
        if index < 1: continue

        if pd.notna(value):
            if 'polyamide' in value or 'CAPROLACTAM' in value:
                tmp_dic = {}
                tmp_dic['name'] = value

                if pd.notna(df.iloc[index, quantity_column_index]):
                    quantity_str = df.iloc[index, quantity_column_index].replace(',', '')
                    tmp_dic['quantity'] = float(quantity_str)
                else:
                    tmp_dic['quantity'] = 0

                if pd.notna(df.iloc[index, amount_column_index]):
                    # 去除千位分隔符
                    amount_str = df.iloc[index, amount_column_index].replace(',', '')
                    tmp_dic['amount'] = float(amount_str)
                else:
                    tmp_dic['amount'] = 0
                
                tmp_dic['tonns_name'] = name_dict[value]

            
                sale_list.append(tmp_dic)
            else:
                other_str = df.iloc[index, amount_column_index].replace(',', '')
                other['amount'] += float(other_str)
        
    sale_list.append(other)

    return sale_list
def handle_account_balance(table, product_list):
    # 读取Excel文件
    df = pd.read_excel(table, sheet_name=0,header=None)

    # 遍历第一行，找到'本期贷方发生金额'所在的列索引
    credit_column_index = None
    borrow_column_index = None
    end_quantity_column_index = None
    end_amount_column_index = None
    for index, value in enumerate(df.iloc[0]):
        if value == '本期贷方发生金额':
            credit_column_index = index
        elif value == '本期借方发生金额':
            borrow_column_index = index
        elif value == '期末余额数量':
            end_quantity_column_index = index
        elif value == '期末余额金额':
            end_amount_column_index = index
    # 遍历表格，提取期末金额和数量
    checked_list = []
    for key in name_dict.keys():
        for index, value in enumerate(df[2]):
            # 排除 非1402和1405开头的行
            filter_v = df.iloc[index, 1]
            if False == pd.notna(filter_v) or ('1402' not in filter_v and '1405' not in filter_v):
                continue

            tmp_dic = {}
            if pd.notna(value) and key in value and 'product' not in value and 'tariff' not in value and 'transfer' not in value and 'Tariff' not in value:
                first = True
                for item in checked_list:
                    if item['name'] == key:
                        first = False

                        tmp_value = df.iloc[index, end_quantity_column_index]
                        if pd.notna(tmp_value):
                            item['quantity'] += float(tmp_value)

                        tmp_value = df.iloc[index, end_amount_column_index]
                        if pd.notna(tmp_value):
                            item['amount'] += float(tmp_value)

                if first:
                    tmp_dic['name'] = key
                    tmp_dic['tonns_name'] = name_dict[key]

                    tmp_value = df.iloc[index, end_quantity_column_index]
                    if pd.notna(tmp_value):
                        tmp_dic['quantity'] = float(tmp_value)
                    else:
                        tmp_dic['quantity'] = 0
                        
                    tmp_value = df.iloc[index, end_amount_column_index]
                    if pd.notna(tmp_value):
                        tmp_dic['amount'] = float(tmp_value)
                    else:
                        tmp_dic['amount'] = 0
                    checked_list.append(tmp_dic)
                
                continue

    # 记录other
    other = {}
    # 记录other的金额
    other_amount = 0

    for item in product_list:
        # 找到other,并记录下来
        if item['name'] == 'other':
            other = item
            continue

        for index, value in enumerate(df[2]):
            # 因为后面会排除非1402的行，所以此逻辑要先执行
            # 找到暂估人民币所在的行
            if pd.notna(value) and '暂估人民币' in value:
                tmp_value = df.iloc[index, credit_column_index]
                if pd.notna(tmp_value):
                    other_amount = float(tmp_value)

            # 排除 非1402开头的行
            if False == pd.notna(df.iloc[index, 1]) or '1402' not in df.iloc[index, 1]:
                continue

            # 遍历product_list，通过item['name']匹配到对应的行
            # 然后找到对应的amount,将amount加到item['amount']中
            if pd.notna(value) and item['name'] in value:
                # 根据列表关系下移两行
                tmp_value = df.iloc[index+2, borrow_column_index]
                if pd.notna(tmp_value):
                    # print('name:{},amount:{},index:{},col:{}'.format(item['name'], tmp_value, index, borrow_column_index))
                    item['amount'] += float(tmp_value)

    other['amount'] += other_amount

    return product_list, checked_list
def handle_tonns_table(tonns_table, product_list, sale_list):
    """
    处理吨位表格函数
    
    该函数用于更新给定表格中的产品吨位数量，并保存为新的Excel文件'tonns.xlsx'。
    
    参数:
    - tonns_table: str，原始吨位表格的文件路径。
    - product_list: list，包含产品信息的列表，每个产品信息为一个字典，必须包含'tonns_name'和'amount'键。
    
    返回:
    无返回值，但会生成一个新的Excel文件'tonns.xlsx'，其中包含了更新后的吨位数量。
    """
    # 读取Excel文件
    tonns = load_workbook(tonns_table)
    # 选择名为'amount'的工作表进行操作
    amount = tonns['amount']
    
    # 遍历产品列表，更新每个产品的amount到Purchase (RMB)表格中
    for item in product_list:
        # 只取行5-28，因为根据上下文，只有这些行包含有效的产品数据
        for row in range(5, 29):
            name_cell = amount.cell(row, 2)
            # 当找到匹配的产品名称时，更新对应的月度数量
            if item['tonns_name'] == str(name_cell.value):
                # 根据当前月份计算列索引，以确保更新正确的月度数据
                cell = amount.cell(row, 3+curr_month)
                cell.value = item['amount']
    
    # 遍历sale_list,更新每个产品sale数据到Sales (tn)表格中
    for item in sale_list:
        # 只取行32-55，因为根据上下文，只有这些行包含有效的产品数据
        for row in range(32, 56):
            name_cell = amount.cell(row, 2)
            # 当找到匹配的产品名称时，更新对应的月度数量
            if item['tonns_name'] == str(name_cell.value):
                # 根据当前月份计算列索引，以确保更新正确的月度数据
                cell = amount.cell(row, 3+curr_month)
                cell.value = item['amount']

    # 选择名为'quantity'的工作表进行操作
    quantity = tonns['quantity']
    # 遍历产品列表，更新每个产品的quantity到Purchase (RMB)表格中
    for item in product_list:
        # 只取行5-27，因为根据上下文，只有这些行包含有效的产品数据
        for row in range(5, 28):
            name_cell = quantity.cell(row, 2)
            # 当找到匹配的产品名称时，更新对应的月度数量
            if item['tonns_name'] == str(name_cell.value):
                # 根据当前月份计算列索引，以确保更新正确的月度数据
                cell = quantity.cell(row, 3+curr_month)
                cell.value = item['quantity']  
    
    # 遍历sale_list,更新每个产品sale数据到Sales (tn)表格中
    for item in sale_list:
        # 只取行32-54，因为根据上下文，只有这些行包含有效的产品数据
        for row in range(32, 55):
            name_cell = quantity.cell(row, 2)
            # 当找到匹配的产品名称时，更新对应的月度数量
            if item['tonns_name'] == str(name_cell.value):
                # 根据当前月份计算列索引，以确保更新正确的月度数据
                cell = quantity.cell(row, 3+curr_month)
                cell.value = item['quantity']  

    # 检查并删除已存在的new_tonns_name文件，避免重复生成
    if os.path.exists(new_tonns_name):
        os.remove(new_tonns_name)
    # 保存更新后的Excel文件
    tonns.save(new_tonns_name)
    
def check_tonns_table(checked_list):
    """
    检查吨位表格函数
    
    该函数用于检查给定表格中的产品吨位数量，并返回一个布尔值，表示是否所有产品吨位数量都正确。
    
    参数:
    - checked_list: list，包含检查信息的列表，每个检查信息为一个字典，必须包含'name'和'amount'键。
    
    返回:
    - bool，如果所有产品吨位数量都正确，则返回True，否则返回False。
    """
    print('#'*20 + 'checked_tonns_table' + '#'*20)
    # 启动Excel后台进程，打开文件并保存
    app = xw.App(visible=False)  # 不显示Excel界面
    wb = app.books.open(new_tonns_name)
    wb.api.Save()  # 强制保存计算结果
    wb.close()
    app.quit()
    ammount = pd.read_excel(new_tonns_name, sheet_name='amount', header=None, engine='openpyxl')
    # 只取行56-82，因为根据上下文，只有这些行包含有效的产品数据
    for row in range(55, 82):# 行号从0开始，所以比excel显示的小1
        # 跳过PA-6
        if row == 56 or row == 57:
            continue

        name_cell = ammount.iloc[row, 1]
        value_cell = ammount.iloc[row, 2+curr_month]
        debug_item_cell = 0

        if pd.notna(value_cell):
            value_cell = float(value_cell)
        else:
            value_cell = 0
        # 保留2位小数
        value_cell = round(value_cell, 2)
        debug_org_cell = value_cell

        for item in checked_list:
            # 当找到匹配的产品名称时，更新对应的月度数量
            if item['tonns_name'] == str(name_cell):
                # 根据当前月份计算列索引，以确保更新正确的月度数据
                value_cell -= round(item['amount'], 2)
                debug_item_cell += round(item['amount'], 2)
        if value_cell != 0:
            print(f"{name_cell}的amount校验有误，amount:{debug_org_cell}，get:{debug_item_cell}，请检查！")

    quantity = pd.read_excel(new_tonns_name, sheet_name='quantity', header=None)
    # 只取行58-81，因为根据上下文，只有这些行包含有效的产品数据
    for row in range(59, 82):
        name_cell = quantity.iloc[row, 1]
        value_cell = quantity.iloc[row, 2+curr_month]
        debug_item_cell = 0

        if pd.notna(value_cell):
            value_cell = float(value_cell)
        else:
            value_cell = 0
        # 保留3位小数
        value_cell = round(value_cell, 3)
        debug_org_cell = value_cell

        for item in checked_list:
            # 当找到匹配的产品名称时，更新对应的月度数量
            if item['tonns_name'] == str(name_cell):
                # 根据当前月份计算列索引，以确保更新正确的月度数据
                value_cell -= round(item['quantity'], 2)
                debug_item_cell += round(item['quantity'], 2)
        if value_cell != 0:
            print(f"{name_cell}的quantity校验有误，quantity:{debug_org_cell}，get:{debug_item_cell}，请检查！")
    # 删除临时文件
    # os.remove(tmp_excel)
    print('#'*20 + 'checked_end' + '#'*20)


def main():
    print(' ')
    print('#'*20 + '开始生成tonns table' + '#'*20)
    # excel所在文件夹路径
    folder_path = "."
    # 出库汇总表
    keyword = "出库汇总表"
    outbound_table = find_excel_files_with_keyword(folder_path, keyword)
    sale_list = handle_outbound_summary(outbound_table)
    #################################################################
    keyword = "在途货物余额表"
    # 在途货物余额表
    tansit_table = find_excel_files_with_keyword(folder_path, keyword)
    # 生成产品字典列表
    product_list = handle_transit_table(tansit_table)
    #################################################################
    # 发生额及余额表
    keyword = "发生额及余额表"
    account_balance_table = find_excel_files_with_keyword(folder_path, keyword)
    # 获取发生额及余额表中的金额加到product_list中
    product_list, checked_list = handle_account_balance(account_balance_table, product_list)
    #################################################################
    # tonns
    keyword = 'tonns of good2'
    tonns_table = find_excel_files_with_keyword(folder_path, keyword)
    product_list = handle_tonns_table(tonns_table, product_list, sale_list)
    # 检查tonns
    check_tonns_table(checked_list)

    while True:
        user_input = input("请输入内容(直接回车执行下一步): ")
        if user_input == '':
            print("你选择了执行下一步。")
            break
        else:
            print("你输入的是:", user_input)

    print('#'*20 + '完成    tonns talbe' + '#'*20)

if __name__ == '__main__':
    main()
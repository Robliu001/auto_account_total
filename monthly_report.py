import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
import glob
from copy import copy  # 导入 copy 函数
import calendar
import xlrd  # 导入 xlrd 库，用于读取 .xls 文件
from dateutil.relativedelta import relativedelta

# excel所在文件夹路径
folder_path = "."  # 当前文件夹
# 日期格式
date_fmt = "%Y%m"
# key是tonns里面的名字，value是report中列名字
tonns_name = {}
tonns_name["25"] = "V 25"
tonns_name['24'] = "V 24"
tonns_name['27'] = "V 27"
tonns_name["32"] = "Other"
tonns_name["34"] = "Other"
tonns_name["36"] = "Other"
tonns_name["40"] = "Other"
tonns_name["24 SD"] = "V 24 SD"
tonns_name["34F"] = "V 34 (F)"
tonns_name["PA-6 recycled"] = "Second PA"
tonns_name["Caprolactam"] = "CPL"
tonns_name["PA-6"] = "PA"
tonns_name["Polyamide light-stabilized yarn 93.5 tex"] = "Yarn"
tonns_name["Polyamide not the  thermostabilized yarn 187tex"] = "Yarn"
tonns_name["yarn 144tex"] = "Yarn"
tonns_name["yarn 94tex"] = "Yarn"
tonns_name["Polyamide fiber 1 tex"] = "Yarn"
tonns_name["Polyamide fiber 0.48 tex"] = "Yarn"
tonns_name["Polyamide fiber 0.68 tex"] = "Yarn"
tonns_name["Polyamide fiber 0.33 tex"] = "Yarn"
tonns_name["NYLON 6 Dipped NTCF"] = "Tyre cord"
# key是发生额及余额表中列名称，value是tonns中列的名称
name_dict = {}
name_dict['polyamide-6(volgamid 25)'] = '25'
name_dict['polyamide-6(volgamid 27)'] = '27'
name_dict['polyamide-6(volgamid 34F)'] = '34F'
name_dict['polyamide-6(volgamid 24)'] = '24'
name_dict['polyamide-6(volgamid 24SD)'] = '24 SD'
name_dict['polyamide-6(volgamid 32)'] = '32'
name_dict['polyamide-6(volgamid 34)'] = '34'
name_dict['polyamide-6(volgamid 36)'] = '36'
name_dict['polyamide-6(volgamid 40)'] = '40'
name_dict['CAPROLACTAM'] = 'Caprolactam'
name_dict['Nylon 6 dipped NTCF'] = 'NYLON 6 Dipped NTCF'
# name_dict['在途物资-materials in transit'] = 'Inventories (tn)'
# name_dict['库存商品-Commodity Stocks'] = 'Inventories (tn)'
# name_dict['其他非主营'] = 'others '

def load_workbook_with_xlrd(path):
    if path.endswith('.xls') or path.endswith('.XLS'):
        workbook = xlrd.open_workbook(path)
        workbook_xlsx = openpyxl.Workbook()
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            ws = workbook_xlsx.create_sheet(title=sheet_name)
            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    ws.cell(row=row_idx + 1, column=col_idx + 1, value=sheet.cell_value(row_idx, col_idx))
        return workbook_xlsx
    else:
        return load_workbook(path)
    
def convert_date_format_en(date_str):
    # 解析原始字符串到datetime对象
    date_obj = datetime.strptime(date_str, date_fmt)
    # 创建新的日期字符串，月份全写，后面跟上年份的后两位
    new_date_str = date_obj.strftime('%B %y')
    
    return new_date_str
def convert_date_format(date_str, split_char):
    # 解析原始字符串到datetime对象
    date_obj = datetime.strptime(date_str, date_fmt)
    # 获取该月的最后一天
    last_day = calendar.monthrange(date_obj.year, date_obj.month)[1]
    # 创建新的日期字符串
    new_date_str = '{:02d}{}{:02d}{}{:02d}'.format(last_day, split_char, date_obj.month, split_char, date_obj.year % 100)
    
    return new_date_str

def find_excel_files_with_keyword(folder_path, keyword):
    # 构建搜索路径，匹配所有 Excel 文件（支持 .xlsx 和 .xls）
    search_pattern = os.path.join(folder_path, f"{keyword}*.xls*")
    # 使用 glob 查找匹配的文件
    matching_files = glob.glob(search_pattern)
    
    return matching_files[0]
def save_excel_file_for_value(path):
    month_wb = load_workbook(path, data_only=True)
    month_wb.save(path)
def copy_example_sheet_add_to_monthly(monthly_path, example_sheet_path):
    monthly_wb = load_workbook(monthly_path)
    example_wb = load_workbook(example_sheet_path)
    sheet_name = monthly_wb.sheetnames[-1]
    last_month_sheet = monthly_wb[sheet_name]
    sheet_name = datetime.strptime(sheet_name, date_fmt)
    # 计算下一个月
    if sheet_name.month < 12:
        next_month_date = sheet_name.replace(month=sheet_name.month + 1)
    else:
        # 如果是12月，则增加到下一年的1月
        next_month_date = sheet_name.replace(year=sheet_name.year + 1, month=1)
    # 新sheet名称
    sheet_name = next_month_date.strftime(date_fmt)
    last_year_date = next_month_date.replace(year=next_month_date.year - 1)
    last_year_sheet = monthly_wb[last_year_date.strftime(date_fmt)]
    ex_sheet = example_wb['example']
    # 如果目标文件中已存在同名工作表，先删除
    if sheet_name in monthly_wb.sheetnames:
        monthly_wb.remove(monthly_wb[sheet_name])
    # 复制工作表到目标工作簿
    new_sheet = monthly_wb.create_sheet(sheet_name)
    # 复制单元格内容和格式
    for row in ex_sheet.iter_rows():
        for cell in row:
            target_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = cell.number_format
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)

    # 复制列宽和行高
    for col in ex_sheet.column_dimensions:
        new_sheet.column_dimensions[col] = copy(ex_sheet.column_dimensions[col])

    for row in ex_sheet.row_dimensions:
        new_sheet.row_dimensions[row] = copy(ex_sheet.row_dimensions[row])
    # 复制合并单元格
    for merged_range in ex_sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_range))

    return monthly_wb, new_sheet, last_month_sheet, last_year_sheet
def write_date_to_product_tbl(new_sheet, last_month_sheet, last_year_sheet):
    # stock compare
    name = convert_date_format(last_month_sheet.title, '.')
    new_sheet['C3'].value = new_sheet['C3'].value + name
    new_sheet['M3'].value = new_sheet['M3'].value + name
    name = convert_date_format(new_sheet.title, '.')
    new_sheet['I3'].value = new_sheet['I3'].value + name
    # 本月
    name = convert_date_format_en(new_sheet.title)
    new_sheet['E3'].value = new_sheet['E3'].value + name
    new_sheet['F3'].value = new_sheet['F3'].value + name
    new_sheet['R3'].value = new_sheet['R3'].value + name
    new_sheet['V3'].value = new_sheet['V3'].value + name
    # 上一月
    name = convert_date_format_en(last_month_sheet.title)
    new_sheet['O3'].value = new_sheet['O3'].value + name
    new_sheet['R4'].value = new_sheet['R4'].value + name
    new_sheet['V4'].value = new_sheet['V4'].value + name
    # 上一年
    name = convert_date_format_en(last_year_sheet.title)
    new_sheet['T3'].value = new_sheet['T3'].value + name
    # 日期
    # Total Delivery
    name = datetime.strptime(last_year_sheet.title, date_fmt)
    new_sheet['Y5'].value = '{} м {}'.format(name.month, name.year)
    # 当前日期，下面计算下一个月还会用到
    name = datetime.strptime(new_sheet.title, date_fmt)
    new_sheet['X5'].value = '{} м {}'.format(name.month, name.year)
    # TOP 10 customers
    new_sheet['B38'] = convert_date_format(new_sheet.title, '/')
    new_sheet['F38'] = convert_date_format(last_month_sheet.title, '/')
    new_sheet['N38'] = convert_date_format(last_year_sheet.title, '/')
    # 计算下一个月
    if name.month < 12:
        name = name.replace(month=name.month + 1)
    else:
        # 如果是12月，则增加到下一年的1月
        name = name.replace(year=name.year + 1, month=1)
    # 新sheet名称
    name = name.strftime(date_fmt)
    name = convert_date_format_en(name)
    new_sheet['AA3'].value = new_sheet['AA3'].value + name
    #  plan 和 fact
    name = convert_date_format_en(new_sheet.title)
    new_sheet['D67'].value = f'Plan {name}'
    new_sheet['E67'].value = f'Fact {name}'
    name = convert_date_format_en(last_year_sheet.title)
    new_sheet['F67'].value = f'Fact {name}'
    new_sheet['G67'].value = new_sheet['E67'].value + ' / ' + new_sheet['D67'].value
    new_sheet['I67'].value = new_sheet['E67'].value + ' / ' + new_sheet['F67'].value

    cur_date = datetime.strptime(new_sheet.title, date_fmt)
    new_sheet['K67'].value = f'Plan {cur_date.month}m.{cur_date.year}'
    new_sheet['L67'].value = f'Fact {cur_date.month}m.{cur_date.year}'
    last_date = datetime.strptime(new_sheet.title, date_fmt)
    new_sheet['M67'].value = f'Fact {last_date.month}m.{last_date.year}'

    new_sheet['N67'].value = new_sheet['L67'].value + ' / ' + new_sheet['K67'].value
    new_sheet['P67'].value = new_sheet['L67'].value + ' / ' + new_sheet['M67'].value
    return

def copy_last_data_to_new(new_sheet, last_month_sheet, last_year_sheet):
    start_row, end_row = 7, 32
    # 复制上月数据
    col = 3 # C列
    # total 列
    for row in range(start_row, end_row + 1):
        # 获取源单元格
        source_cell = last_month_sheet.cell(row=row, column=col+6)
        # 获取目标单元格
        target_cell = new_sheet.cell(row=row, column=col, value=source_cell.value)
        target_cell.fill = copy(source_cell.fill)

    col = 4 # C列
    # Incl on the way 列
    for row in range(start_row, end_row + 1):
        # 获取源单元格
        source_cell = last_month_sheet.cell(row=row, column=col+8)
        # 获取目标单元格
        target_cell = new_sheet.cell(row=row, column=col, value=source_cell.value)
        target_cell.fill = copy(source_cell.fill)
    # 复制上月Delivery
    start_col, end_col = 6, 8  # F列到H列

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            # 获取源单元格
            source_cell = last_month_sheet.cell(row=row, column=col)
            # 获取目标单元格
            target_cell = new_sheet.cell(row=row, column=col+9, value=source_cell.value)
            target_cell.fill = copy(source_cell.fill)
    # 复制上一年Delivery
    col = 6 # G列
    # total 列
    for row in range(start_row, end_row + 1):
        # 获取源单元格
        source_cell = last_year_sheet.cell(row=row, column=col)
        # 获取目标单元格
        target_cell = new_sheet.cell(row=row, column=col+14, value=source_cell.value)
        target_cell.fill = copy(source_cell.fill)

    col = 8 # I列
    # Incl on the way 列
    for row in range(start_row, end_row + 1):
        # 获取源单元格
        source_cell = last_year_sheet.cell(row=row, column=col)
        # 获取目标单元格
        target_cell = new_sheet.cell(row=row, column=col+13, value=source_cell.value)
        target_cell.fill = copy(source_cell.fill)
    # 复制上一年的total delivery
    col = 24 # I列
    # Incl on the way 列
    for row in range(start_row, end_row + 1):
        # 获取源单元格
        source_cell = last_year_sheet.cell(row=row, column=col)
        # 获取目标单元格
        target_cell = new_sheet.cell(row=row, column=col+1, value=source_cell.value)
        target_cell.fill = copy(source_cell.fill)
    # TOP 10 customers
    # 上个月的TOP 10 customers
    start_row = 41
    end_row = 53
    start_col = 2
    end_col = 5
    offset = 4
    for col in range(start_col, end_col + 1):
        for row in range(start_row, end_row + 1):
            # 跳过合并单元格
            if (row == 52 or row == 53) and ((col+4) == 7):
                continue
            # 获取源单元格
            source_cell = last_month_sheet.cell(row=row, column=col)
            # 获取目标单元格
            # H列是合并单元格，跳过
            if (col+offset) == 8:
                offset = 5
            target_cell = new_sheet.cell(row=row, column=col+offset, value=source_cell.value)
            target_cell.fill = copy(source_cell.fill)
    
    # 上一年的TOP 10 customers
    start_col = 2
    end_col = 3
    for col in range(start_col, end_col + 1):
        for row in range(start_row, end_row + 1):
            # 跳过合并单元格
            if (row == 52 or row == 53) and ((col+12) == 15):
                continue
            # 获取源单元格
            source_cell = last_year_sheet.cell(row=row, column=col)
            # 获取目标单元格
            target_cell = new_sheet.cell(row=row, column=col+12, value=source_cell.value)
            target_cell.fill = copy(source_cell.fill)

    # 复制上月的TOP 5 suppliers
    start_row = 55
    end_row = 59
    start_col = 3
    end_col = 5
    for col in range(start_col, end_col + 1):
        for row in range(start_row, end_row + 1):
            # 获取源单元格
            source_cell = last_month_sheet.cell(row=row, column=col)
            # 获取目标单元格
            target_cell = new_sheet.cell(row=row, column=col+5, value=source_cell.value)
            target_cell.fill = copy(source_cell.fill)
    
    # 复制上一年的TOP 5 suppliers
    # 仅需要total数据
    for row in range(start_row, end_row + 1):
        # 获取源单元格
        source_cell = last_year_sheet.cell(row=row, column=start_col)
        # 获取目标单元格
        target_cell = new_sheet.cell(row=row, column=start_col+12, value=source_cell.value)
        target_cell.fill = copy(source_cell.fill)

    # 上年同期实际发生
    start_row = 69
    end_row = 81
    # 同期fact
    start_col = 5
    # 合计fact
    start_total_col = 11
    for row in range(start_row, end_row + 1):
         # 获取源单元格
        source_cell = last_year_sheet.cell(row=row, column=start_col)
        # 获取目标单元格
        target_cell = new_sheet.cell(row=row, column=start_col+1, value=source_cell.value)
        target_cell.fill = copy(source_cell.fill)

        # 上年同期合计实际发生
        # 获取源单元格
        source_cell = last_year_sheet.cell(row=row, column=start_total_col)
        # 获取目标单元格
        target_cell = new_sheet.cell(row=row, column=start_total_col+2, value=source_cell.value)
        target_cell.fill = copy(source_cell.fill)

def copy_tonns_data_to_report(tonns_path, new_sheet):
    quantity = load_workbook(tonns_path)['quantity']
    amount = load_workbook(tonns_path)['amount']
    # 先获取purchase信息
    purchase = []
    # 解析原始字符串到datetime对象
    date_obj = datetime.strptime(new_sheet.title, date_fmt)
    col = date_obj.month + 3

    for i in range(5, 29):
        tmp = {}
        name = quantity.cell(row=i, column=2).value
        
        for key, value in tonns_name.items():
            if str(name) == key:
                exist = False
                for item in purchase:
                    if item['name'] == value:
                        tmp = item
                        exist = True
                if not exist:
                    tmp['name'] = value
                    tmp['quantity'] = 0
                    tmp['amount'] = 0
                    purchase.append(tmp)

                # quantity
                val = quantity.cell(row=i, column=col).value
                if val is not None:
                    tmp['quantity'] += float(val)
                # amount
                val = amount.cell(row=i, column=col).value
                if val is not None:
                    tmp['amount'] += float(val)/1000
                break
    
    # print(purchase)
    # 获取sales信息
    sale = []
    for i in range(32, 56):
        tmp = {}
        name = quantity.cell(row=i, column=2).value

        for key, value in tonns_name.items():
            if str(name) == key:
                exist = False
                for item in sale:
                    if item['name'] == value:
                        tmp = item
                        exist = True
                if not exist:
                    tmp['name'] = value
                    tmp['quantity'] = 0
                    tmp['amount'] = 0
                    sale.append(tmp)

                # quantity
                val = quantity.cell(row=i, column=col).value
                if val is not None:
                    tmp['quantity'] += float(val)
                # amount
                val = amount.cell(row=i, column=col).value
                if val is not None:
                    tmp['amount'] += float(val)/1000
                break

    # print(sale)
    # 填充数据到report sheet's Receipt(E列) Delivery(F列)
    start_row, end_row = 7, 32
    col_re = 5 # E列
    col_dl = 6 # F列
    for row in range(start_row, end_row + 1):
        # 获取源单元格
        name = str(new_sheet.cell(row=row, column=1).value)# 字符串匹配

        for item in purchase:
            if name == item['name']:
                tmp_amount = item['amount']
                if name == "CPL" or name == "Yarn":
                    tmp_amount = tmp_amount / new_sheet.cell(row=2, column=2).value
                mt_cell = new_sheet.cell(row=row, column=col_re, value=item['quantity'])
                qian_cell = new_sheet.cell(row=row+1, column=col_re, value=tmp_amount)
                break
        
        for item in sale:
            if name == item['name']:
                tmp_amount = item['amount']
                if name == "CPL" or name == "Yarn":
                    tmp_amount = tmp_amount / new_sheet.cell(row=2, column=2).value         
                mt_cell = new_sheet.cell(row=row, column=col_dl, value=item['quantity'])
                qian_cell = new_sheet.cell(row=row+1, column=col_dl, value=tmp_amount)
                break
                
def copy_transit_data_to_report(transit_path, new_sheet):
    # 读取Excel文件
    transit = pd.read_excel(transit_path, sheet_name='在途货物余额表', header=None, engine='openpyxl')
    product_list = []
    other = {}
    other['name'] = 'Other'
    other['quantity'] = 0
    other['amount'] = 0    
    # 遍历表格，提取产品信息
    for index, value in enumerate(transit[1]):
        if index < 5: continue  

        if pd.notna(value):
            exist = False

            for key in name_dict.keys():
                if key in str(value):
                    exist = True
                    break

            if True == exist:
                tmp_dic = {}
                name = tonns_name[name_dict[key]]
                tmp_dic['name'] = name

                if pd.notna(transit.iloc[index, 14]):
                    tmp_dic['quantity'] = float(transit.iloc[index, 14])
                else:
                    tmp_dic['quantity'] = 0

                if pd.notna(transit.iloc[index, 8]):
                    tmp_dic['amount'] = float(transit.iloc[index, 15])/1000
                else:
                    tmp_dic['amount'] = 0
                # 添加到product_list
                product_list.append(tmp_dic)
            else:
                other['amount'] += float(transit.iloc[index, 15])/1000
    # 添加other
    product_list.append(other)
    # print(product_list)
    # 填充数据到report sheet's Incl on the way(L列)   
    start_row, end_row = 7, 32
    col_in = 12 # L列
    for row in range(start_row, end_row + 1):
        # 获取源单元格
        name = str(new_sheet.cell(row=row, column=1).value)# 字符串匹配

        for item in product_list:
            if name == item['name']:
                mt_cell = new_sheet.cell(row=row, column=col_in, value=item['quantity'])
                qian_cell = new_sheet.cell(row=row+1, column=col_in, value=item['amount'])
                break

def create_current_total_delivery(new_sheet, last_month_sheet):
    delivery = new_sheet
    # 当前日期，下面计算下一个月还会用到
    current_date = datetime.strptime(delivery.title, date_fmt)
    # 当年所有月份的数据的总和
    # 填充数据到report sheet's Total Delivery(X列)
    start_row, end_row = 7, 32
    col_de = 24 # X列, openpyxl, A列序号是1
    col_ct = 6 # F列

    for row in range(start_row, end_row + 1):
        if current_date.month != 1:
            delivery.cell(row=row, column=col_de).value = last_month_sheet.cell(row=row, column=col_de).value + delivery.cell(row=row, column=col_ct).value
        else:
            delivery.cell(row=row, column=col_de).value = delivery.cell(row=row, column=col_ct).value

def create_top_10_customer_table(new_sheet, last_month_sheet, last_year_sheet, receivable_path, customer_sheet):
    customer_dict = []
    receive_wb = load_workbook_with_xlrd(receivable_path)
    total_sheet = receive_wb['汇总']
    other_dict = {
        'code': '0000',
        'func_curr_balan': 0,
        'over_due': 0,
        '3month': 0,
        'simple_name': 'Other customers'
    }
    # 获取所有客户的汇总信息,忽略最后两行
    for row in range(2, total_sheet.max_row + 1 - 2):
        customer_code = total_sheet.cell(row=row, column=1).value
        if customer_code is not None:
            func_curr_balan = total_sheet.cell(row=row, column=3).value / 1000
            func_curr_balan_bill = total_sheet.cell(row=row, column=4).value / 1000
            over_due = func_curr_balan - func_curr_balan_bill

            if len(customer_dict) < 10:
                # 如果customer_dict中少于10个客户，直接插入并排序
                tmp_item = {
                    'code': customer_code,
                    'func_curr_balan': func_curr_balan,
                    'over_due': over_due,
                    '3month': 0,
                    'simple_name': ''
                }
                customer_dict.append(tmp_item)
                customer_dict.sort(key=lambda x: x['func_curr_balan'], reverse=True)
            else:
                # 如果customer_dict中已有10个客户，比较func_curr_balan
                last_balance = customer_dict[-1]['func_curr_balan']
                # 如果新客户的func_curr_balan大于最后一个客户的func_curr_balan，则替换最后一个客户
                if func_curr_balan > last_balance:
                     # last_balance累加到'Other customers'
                    other_dict['func_curr_balan'] += last_balance
                    other_dict['over_due'] += customer_dict[-1]['over_due']              
                    tmp_item = {
                        'code': customer_code,
                        'func_curr_balan': func_curr_balan,
                        'over_due': over_due,
                        '3month': 0,
                        'simple_name': ''
                    }
                    customer_dict[-1] = tmp_item
                    customer_dict.sort(key=lambda x: x['func_curr_balan'], reverse=True)
                else:
                    # 否则，将func_curr_balan累加到'Other customers'
                    other_dict['func_curr_balan'] += func_curr_balan
                    other_dict['over_due'] += over_due
    # 将'Other customers'添加到customer_dict中
    customer_dict.append(other_dict)
    # 创建一个临时列表，用检索，添加simple_name
    tmp_list = customer_dict.copy()
    for row in range(2, customer_sheet.max_row + 1):
        customer_code = customer_sheet.cell(row=row, column=3).value
        # 遍历tmp_list找到对应的客户名称
        for item in tmp_list:
            if customer_code == item['code']:
                # 获取客户简称
                customer_name = customer_sheet.cell(row=row, column=5).value
                # 将客户名称添加到字典中
                item['simple_name'] = customer_name
                # 从tmp_list中删除该客户，减少循环遍历的次数
                tmp_list.remove(item)
                break
        # 如果匹配完成，则跳出循环
        if len(tmp_list) == 0:
            break

    # 找出超期三个月的客户
    date_obj = datetime.strptime(new_sheet.title, date_fmt)
    # 日期改为该月的最后一天
    day_tmp = calendar.monthrange(date_obj.year, date_obj.month)[1]
    date_obj = date_obj.replace(day=day_tmp)
    # 计算三个月前的那一天日期
    three_months_ago = date_obj - relativedelta(months=3)
    # 遍历‘明细’sheet，获取超期三个月的客户
    detail_sheet = receive_wb['明细']
    for row in range(2, detail_sheet.max_row + 1):
        # 获取原币余额
        ybye = detail_sheet.cell(row=row, column=17).value
        if ybye is not None and float(ybye) > 0:
            # 获取客户的到期日期
            due_date = detail_sheet.cell(row=row, column=10).value
            # 如果到期日期在三个月之前，则将func_curr_balan累加到'Other customers'
            if isinstance(due_date, str):
                due_date = datetime.strptime(due_date, '%Y-%m-%d')
            if due_date < three_months_ago:
                customer_code = detail_sheet.cell(row=row, column=1).value
                find_r = False
                yuee = detail_sheet.cell(row=row, column=18).value / 1000
                for item in customer_dict:
                    if customer_code == item['code']:
                        item['3month'] += yuee
                        find_r = True
                        break
                # 如果没有找到对应的客户，则将yuee累加到'Other customers'
                if not find_r:
                    other_dict['3month'] += yuee

    # 打印结果（可选）
    # for item in customer_dict:
    #     print(item)

    start_row = 41
    end_row = 50
    last_month_dict = []
    last_year_dict = []
    for row in range(start_row, end_row+1):
        # 上一月
        name = last_month_sheet.cell(row=row, column=2).value
        if isinstance(name, str):
            tmp_item = {
                'code': '',
                'func_curr_balan': last_month_sheet.cell(row=row, column=3).value,
                'over_due': last_month_sheet.cell(row=row, column=5).value,
                '3month': last_month_sheet.cell(row=row, column=4).value,
                'simple_name': name
            }    
            last_month_dict.append(tmp_item)
        # 上一年
        name = last_year_sheet.cell(row=row, column=2).value
        if isinstance(name, str):
            tmp_item = {
                'code': '',
                'func_curr_balan': last_year_sheet.cell(row=row, column=3).value,
                'over_due': last_year_sheet.cell(row=row, column=5).value,
                '3month': last_year_sheet.cell(row=row, column=4).value,
                'simple_name': name
            }
            last_year_dict.append(tmp_item)
    # 向new_sheet写入本月top 10客户数据
    for i, item in enumerate(customer_dict):
        if 10 == i:
            break
        new_sheet.cell(row=i+start_row, column=2).value = f"{str(i + 1)}.{item['simple_name']}"
        new_sheet.cell(row=i+start_row, column=3).value = item['func_curr_balan']
        new_sheet.cell(row=i+start_row, column=4).value = item['3month']
        new_sheet.cell(row=i+start_row, column=5).value = item['over_due']
        # 与上一月的数据进行比较
        find_r = False
        for last_item in last_month_dict:
            if item['simple_name'] in last_item['simple_name']:
                new_sheet.cell(row=i+start_row, column=11).value = item['func_curr_balan'] - last_item['func_curr_balan']
                new_sheet.cell(row=i+start_row, column=12).value = item['3month'] - last_item['3month']
                new_sheet.cell(row=i+start_row, column=13).value = item['over_due'] - last_item['over_due']
                find_r = True
                break
        if not find_r:
            new_sheet.cell(row=i+start_row, column=11).value = item['func_curr_balan']
            new_sheet.cell(row=i+start_row, column=12).value = item['3month']
            new_sheet.cell(row=i+start_row, column=13).value = item['over_due']
        # 与上一年的数据进行比较
        find_r = False
        for last_item in last_year_dict:
            if item['simple_name'] in last_item['simple_name']:
                new_sheet.cell(row=i+start_row, column=16).value = item['func_curr_balan'] - last_item['func_curr_balan']
                # last_item['func_curr_balan']应该不能等于0
                new_sheet.cell(row=i+start_row, column=17).value = item['func_curr_balan'] / last_item['func_curr_balan']
                find_r = True
                break
        if not find_r:
            new_sheet.cell(row=i+start_row, column=16).value = item['func_curr_balan']

    # other customers数据
    new_sheet.cell(row=start_row+11, column=2).value = customer_dict[-1]['func_curr_balan']
    new_sheet.cell(row=start_row+11, column=4).value = customer_dict[-1]['3month']
    new_sheet.cell(row=start_row+11, column=5).value = customer_dict[-1]['over_due']

def create_top_5_supplier_table(new_sheet, last_month_sheet, last_year_sheet, supplier_path, supplier_sheet):
    # 先遍历供应商表，创建供应商字典
    supplier_dict = []
    for row in range(2, supplier_sheet.max_row):
        tmp_dic = {
            'code': supplier_sheet.cell(row=row, column=1).value,
            'simple_name': supplier_sheet.cell(row=row, column=2).value,
            'func_curr_balan': 0,
            'over_due': 0,
            '3month': 0,
        }

        supplier_dict.append(tmp_dic)
    # 先不添加到supplier_dict中，等到最后再添加
    other_dict = {
        'code': '0000',
        'func_curr_balan': 0,
        'over_due': 0,
        '3month': 0,
        'simple_name': 'warehouse fee and transportation fee include the purchase in china'
    }
    supplier_wb = load_workbook_with_xlrd(supplier_path)
    total_sheet = supplier_wb['汇总']
    # 获取所有客户的汇总信息,忽略最后两行
    for row in range(2, total_sheet.max_row + 1 - 2):
        supplier_code = total_sheet.cell(row=row, column=1).value
        if supplier_code is not None:
            func_curr_balan = total_sheet.cell(row=row, column=3).value
            
            if func_curr_balan is None or func_curr_balan < 0:
                continue
            else:
                func_curr_balan = func_curr_balan / 1000

            func_curr_balan_bill = total_sheet.cell(row=row, column=4).value / 1000
            over_due = func_curr_balan - func_curr_balan_bill
            find_r = False

            for item in supplier_dict:
                if item['code'] == supplier_code:
                    item['func_curr_balan'] += func_curr_balan
                    item['over_due'] += over_due
                    find_r = True
                    break
            if not find_r:
                other_dict['func_curr_balan'] += func_curr_balan
                other_dict['over_due'] += over_due
    # 找出超期三个月的客户
    date_obj = datetime.strptime(new_sheet.title, date_fmt)
    # 日期改为该月的最后一天
    day_tmp = calendar.monthrange(date_obj.year, date_obj.month)[1]
    date_obj = date_obj.replace(day=day_tmp)
    # 计算三个月前的那一天日期
    three_months_ago = date_obj - relativedelta(months=3)
    # 遍历‘明细’sheet，获取超期三个月的客户
    detail_sheet = supplier_wb['明细']
    for row in range(2, detail_sheet.max_row + 1):
        # 获取原币余额
        ybye = detail_sheet.cell(row=row, column=16).value
        if ybye is not None and float(ybye) > 0:
            # 获取客户的到期日期
            due_date = detail_sheet.cell(row=row, column=10).value
            # 如果到期日期在三个月之前，则将func_curr_balan累加到'Other customers'
            if isinstance(due_date, str):
                due_date = datetime.strptime(due_date, '%Y-%m-%d')
            if due_date < three_months_ago:
                customer_code = detail_sheet.cell(row=row, column=1).value
                find_r = False
                yuee = detail_sheet.cell(row=row, column=18).value / 1000
                for item in supplier_dict:
                    if customer_code == item['code']:
                        item['3month'] += yuee
                        find_r = True
                        break
                # 如果没有找到对应的客户，则将yuee累加到'Other customers'
                if not find_r:
                    other_dict['3month'] += yuee
    
    supplier_dict.append(other_dict)
    # 发生额及余额表的应付暂估 220203和220204
    keyword = "发生额及余额表"
    account_balance_table = find_excel_files_with_keyword(folder_path, keyword)
    account_wb = load_workbook_with_xlrd(account_balance_table)
    account_sheet = account_wb['sheet1']
    # 遍历account_sheet，获取应付暂估值
    end_banlance_col = 0
    for col in range(2, account_sheet.max_column + 1):
        if '期末余额金额' in account_sheet.cell(row=1, column=col).value:
            end_banlance_col = col
    
    assert end_banlance_col != 0, '没有找到期末余额金额列'

    for row in range(2, account_sheet.max_row + 1 - 2):
        tmp_code = account_sheet.cell(row=row, column=2).value
        if '220203' in tmp_code or '220204' in tmp_code:
            # 获取应付暂估值
            func_curr_balan = account_sheet.cell(row=row, column=end_banlance_col).value
            
            if func_curr_balan is None or func_curr_balan < 0:
                continue
            else:
                func_curr_balan = func_curr_balan / 1000
                other_dict['func_curr_balan'] += func_curr_balan
    
    # 写入数据到report sheet's TOP 5 suppliers
    start_row = 55
    start_col = 3
    for i, item in enumerate(supplier_dict):
        new_sheet.cell(row=i+start_row, column=start_col).value = item['func_curr_balan']
        new_sheet.cell(row=i+start_row, column=start_col+1).value = item['3month']
        new_sheet.cell(row=i+start_row, column=start_col+2).value = item['over_due']

def search_data_from(sheet, expect_name, match_col, data_col):
    result = 0

    for row in range(1, sheet.max_row + 1):
        name = sheet.cell(row=row, column=match_col).value

        if name is not None and name == expect_name:
            result = sheet.cell(row=row, column=data_col).value

            if result is not None and (type(result) == float or type(result) == int):
                return result
            else:
                return 0
            
    return result
def create_fact_and_plan_table(new_sheet, last_month_sheet, last_year_sheet):
    fact_plan_dic = {}
    start_row = 69
    end_row = 81
    # 创建fact_plan_dic
    for row in range(start_row, end_row + 1):
        item = {}
        item['name'] = new_sheet.cell(row=row, column=1).value
        item['plan'] = new_sheet.cell(row=row, column=4).value
        item['match_rule'] = ''#记录匹配规则，仅记录作用
        fact_plan_dic[item['name']] = item
        item['fact'] = 0

    # 本月fact数据从发生额及余额表中获取
    keyword = "发生额及余额表"
    account_balance_table = find_excel_files_with_keyword(folder_path, keyword)
    account_wb = load_workbook_with_xlrd(account_balance_table)
    account_sheet = account_wb['sheet1']
    # Revenues from sale of goods: =SUM(E70:E72)
    tmp_total = 0
    item = fact_plan_dic['CPL']
    item['match_rule'] = "600104"
    item['fact'] = search_data_from(account_sheet, item['match_rule'], 2, 8)
    tmp_total += item['fact']

    item = fact_plan_dic['PA6']
    item['match_rule'] = "600103"
    item['fact'] = search_data_from(account_sheet, item['match_rule'], 2, 8)
    tmp_total += item['fact']

    item = fact_plan_dic['Others']
    item['match_rule'] = "6001" # '6001-'600103-'600104
    item['fact'] = search_data_from(account_sheet, item['match_rule'], 2, 8)
    item['fact'] = item['fact'] - fact_plan_dic['CPL']['fact'] - fact_plan_dic['PA6']['fact']
    tmp_total += item['fact']

    item = fact_plan_dic['Revenues from sale of goods']
    item['fact'] = tmp_total
    # Expenses for the sale of goods incl: 是=SUM(E74:E79)
    tmp_total = 0
    item = fact_plan_dic['Cost of goods']
    item['match_rule'] = "6401"
    item['fact'] = search_data_from(account_sheet, item['match_rule'], 2, 8)
    tmp_total += item['fact']

    item = fact_plan_dic['Sales taxes']
    item['match_rule'] = "6403"
    item['fact'] = search_data_from(account_sheet, item['match_rule'], 2, 8)
    tmp_total += item['fact']

    item = fact_plan_dic['Operating costs']
    item['match_rule'] = "6601"
    item['fact'] = search_data_from(account_sheet, item['match_rule'], 2, 8)
    tmp_total += item['fact']

    item = fact_plan_dic['Personnel costs']
    item['match_rule'] = "660201" # 660201-6602010401-6602010402-6602011004
    item['fact'] = search_data_from(account_sheet, item['match_rule'], 2, 8)
    item['fact'] -= search_data_from(account_sheet,"6602010401" , 2, 8)
    item['fact'] -= search_data_from(account_sheet,"6602010402" , 2, 8)
    item['fact'] -= search_data_from(account_sheet,"6602011004" , 2, 8)
    tmp_total += item['fact']

    pers_v = item['fact']
    item = fact_plan_dic['Other administrative expenses']
    item['match_rule'] = "6602" # =6602-660224-660219-660205-pers_v
    item['fact'] = search_data_from(account_sheet, item['match_rule'], 2, 8)
    item['fact'] -= search_data_from(account_sheet, "660224", 2, 8)
    item['fact'] -= search_data_from(account_sheet, "660219", 2, 8)
    item['fact'] -= search_data_from(account_sheet, "660205", 2, 8)
    item['fact'] -= pers_v
    tmp_total += item['fact']

    item = fact_plan_dic['Financial expenses']
    item['match_rule'] = "660302" # =660302+660303
    item['fact'] = search_data_from(account_sheet, item['match_rule'], 2, 8)
    item['fact'] += search_data_from(account_sheet, "660303", 2, 8)
    tmp_total += item['fact']

    item = fact_plan_dic['Expenses for the sale of goods incl:']
    item['fact'] = tmp_total

    item = fact_plan_dic['Other income']
    item['match_rule'] = "660301" # =-660301-660304-6711-660219-660224-660205-+6301+6051
    item['fact'] = 0 - search_data_from(account_sheet, item['match_rule'], 2, 8)
    item['fact'] -= search_data_from(account_sheet, "660304", 2, 8)
    item['fact'] -= search_data_from(account_sheet, "6711", 2, 8)
    item['fact'] -= search_data_from(account_sheet, "660219", 2, 8)
    item['fact'] -= search_data_from(account_sheet, "660224", 2, 8)
    item['fact'] -= search_data_from(account_sheet, "660205", 2, 8)
    item['fact'] += search_data_from(account_sheet, "6301", 2, 8)
    item['fact'] += search_data_from(account_sheet, "6051", 2, 8)

    item = fact_plan_dic['Profit before tax']
    item['match_rule'] = "6801"
    item['fact'] = search_data_from(account_sheet, item['match_rule'], 2, 8)

    
    start_row = 69
    end_row = 81
    dst_col = 5
    cur_dt = datetime.strptime(new_sheet.title, date_fmt)
    i = 0

    for key, value in fact_plan_dic.items():
        # 跳过公式
        if 0 != i and 4 != i:
            new_sheet.cell(row=i+start_row, column=dst_col).value = value['fact']
            # 当前日期
            last_month_plan = 0
            last_month_fact = 0
            # 上月数据
            if cur_dt.month != 1:
                last_month_plan = last_month_sheet.cell(row=i+start_row, column=11).value
                last_month_fact = last_month_sheet.cell(row=i+start_row, column=12).value

            new_sheet.cell(row=i+start_row, column=11).value = value['plan'] + last_month_plan
            new_sheet.cell(row=i+start_row, column=12).value = value['fact'] + last_month_fact
        i += 1

def main():
    print(' ')
    print('#'*20 + '开始生成report table' + '#'*20)
    # 出库汇总表
    keyword = "new monthly report"
    monthly_path = find_excel_files_with_keyword(folder_path, keyword)
    # 将公式变为数值保存
    save_excel_file_for_value(monthly_path)
    # 复制模板，添加新sheet到report,返回excel，新日期sheet，相对新日期的上个月sheet，上一年sheet
    monthly_wb, new_sheet, last_month_sheet, last_year_sheet = copy_example_sheet_add_to_monthly(monthly_path, 'example.xlsx')
    # 完成product表头
    write_date_to_product_tbl(new_sheet, last_month_sheet, last_year_sheet)
    # 复制上月和上年数据
    copy_last_data_to_new(new_sheet, last_month_sheet, last_year_sheet)
    # 复制tonns数据
    tonns_path = find_excel_files_with_keyword(folder_path, "tonns of goods")
    copy_tonns_data_to_report(tonns_path, new_sheet)
    # 复制在途货物余额表
    keyword = "在途货物余额表"
    transit_path = find_excel_files_with_keyword(folder_path, keyword)
    copy_transit_data_to_report(transit_path, new_sheet)
    # 生成total delivery数据
    create_current_total_delivery(new_sheet, last_month_sheet)
    # 生成top 10 customer数据
    keyword = "应收账款"
    receiveable_path = find_excel_files_with_keyword(folder_path, keyword)
    customer_supply = load_workbook('供应商和客户名字.xlsx')
    customer_sheet = customer_supply['客户']
    # 读取客户信息表
    create_top_10_customer_table(new_sheet, last_month_sheet, last_year_sheet, receiveable_path, customer_sheet)
    # 生成top 5 supplier数据
    keyword = "应付账款"
    supplier_path = find_excel_files_with_keyword(folder_path, keyword)
    supplier_sheet = customer_supply['供应商']
    # 读取供应商信息表
    create_top_5_supplier_table(new_sheet, last_month_sheet, last_year_sheet, supplier_path, supplier_sheet)
    # 生成fact and plan数据
    create_fact_and_plan_table(new_sheet, last_month_sheet, last_year_sheet)
    # 保存到新的excel文件
    file_name = 'new monthly report{}.xlsx'.format(new_sheet.title)
    if os.path.exists(file_name):  # 如果文件存在
        # 删除文件
        os.remove(file_name)

    monthly_wb.active = new_sheet
    monthly_wb.save(file_name)

    print('#'*20 + '完成    report table' + '#'*20)

if __name__ == '__main__':
    main()